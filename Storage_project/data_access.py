import json
from openpyxl import Workbook
from typing import TypeAlias, no_type_check
from interfaces import StorageProto, CategoriesProto, OrdersProto
from DTO import (CategoryDTO,
                 StorageDTO,
                 OrderDTO,
                 CategoryStatisticDTO,
                 CommonInfoDTO)


StorageInfo: TypeAlias = list[StorageDTO]
OrderInfo: TypeAlias = list[OrderDTO]
CategoryInfo: TypeAlias = list[CategoryDTO]


class StorageJson(StorageProto):
    def __init__(self, file_name: str) -> None:
        self.file_name = file_name

    def get_storage_info(self) -> StorageInfo:
        with open(self.file_name, encoding='utf8') as file:
            storage = json.load(file)
            res = []
            for row in storage:
                res.append(StorageDTO(name=row["name"],
                                      category=row["category"],
                                      created_at=row["created_at"],
                                      price=row["price"],
                                      updated_at=row["updated_at"],
                                      quantity=row["quantity"],
                                      product_id=row["id"],
                                      add_param=row["add_param"]))
        return res

    def record_with_new_info(self, new_info: StorageInfo) -> None:
        res = []
        for row in new_info:
            res.append({"name": row.name,
                        "category": row.category,
                        "created_at": row.created_at,
                        "price": row.price,
                        "updated_at": row.updated_at,
                        "quantity": row.quantity,
                        "id": row.product_id,
                        "add_param": row.add_param})
        with open(self.file_name, 'w', encoding='utf8') as file:
            json.dump(res, file, indent=2, sort_keys=True)


class CategoriesJson(CategoriesProto):
    def __init__(self, file_name: str) -> None:
        self.file_name = file_name

    def get_all_categories(self) -> CategoryInfo:
        with open(self.file_name, encoding='utf8') as file:
            res = []
            for row in json.load(file):
                res.append(CategoryDTO(category=row['category'],
                                       parameters=row['parameters']))
            return res

    def record_new_category(self,
                            category: str,
                            parameters: list[str]) -> None:
        new_info = self.get_all_categories()
        new_info.append(CategoryDTO(category=category, parameters=parameters))
        res = []
        for row in new_info:
            res.append({"category": row.category,
                        "parameters": row.parameters})
        with open(self.file_name, 'w', encoding='utf8') as file:
            json.dump(res, file, indent=2, sort_keys=True)


class OrdersJson(OrdersProto):
    def __init__(self, file_name: str) -> None:
        self.file_name = file_name

    def get_order_info(self) -> OrderInfo:
        with open(self.file_name, encoding='utf8') as file:
            res = []
            for row in json.load(file):
                res.append(OrderDTO(order_price=row["order price"],
                                    order_id=row['order id'],
                                    goods=row['goods'],
                                    created_at=row['created_at']))
            return res

    def record_order(self,
                     order: OrderDTO) -> None:
        if self.read_file():
            storage = self.get_order_info()
        else:
            storage = []
        with open(self.file_name, 'w', encoding='utf8') as file:
            storage.append(order)
            res = []
            for row in storage:
                res.append({"order price": row.order_price,
                            "order id": row.order_id,
                            "goods": row.goods,
                            "created_at": row.created_at})
            json.dump(res, file, indent=2, sort_keys=True)

    def read_file(self) -> bool:
        with open(self.file_name, encoding='utf8') as file:
            reader = file.read()
            if reader:
                return True
            else:
                return False


@no_type_check
def append_to_ws(worksheet, data_list):
    worksheet.append(data_list)
    return worksheet


def record_to_exel_sheets(info1: 'CategoryStatisticDTO',
                          info2: dict[tuple[str, str, str], int],
                          info3: OrderInfo,
                          info4: 'CommonInfoDTO') -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Sheet1'
    ws1.append(["Category"] + [str(info1.category)])
    quantity_list = [str(item) for item in info1.quantity]
    ws1.append(["Quantity"] + quantity_list)
    revenues_list = [str(item) for item in info1.total_revenues]
    ws1.append(["Total revenues"] + revenues_list)

    wb.create_sheet("Sheet2")
    ws2 = wb["Sheet2"]
    ws2 = append_to_ws(ws2, ['Product name', 'Quantity of sold goods'])
    for key, value in info2.items():
        ws2 = append_to_ws(ws2, [key[0], value])

    wb.create_sheet("Sheet3")
    ws3 = wb["Sheet3"]
    ws3 = append_to_ws(ws3, ["Order id", "Order price", "Creation date",
                             "Goods id", "Goods quantity"])
    for order in info3:
        goods_id = ''
        goods_quantity = ''
        for item in order.goods:
            goods_id += f'{item["product_id"]} '
            goods_quantity += f'{item["quantity"]} '
        result = [order.order_id, order.order_price,
                  order.created_at, goods_id, goods_quantity]
        ws3 = append_to_ws(ws3, result)

    wb.create_sheet("Sheet4")
    ws4 = wb["Sheet4"]
    ws4 = append_to_ws(ws4, ["Total revenues", str(info4.total_revenues)])
    ws4 = append_to_ws(ws4, ["Total amount", str(info4.total_amount)])
    ws4 = append_to_ws(ws4, ["The most popular category",
                             info4.popular_category])
    ws4 = append_to_ws(ws4, ["The most popular product",
                             info4.popular_product])

    wb.save('statistic.xlsx')
